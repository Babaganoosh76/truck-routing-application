# import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
from dateutil import parser
import os
from .load import *

thin_border = Side(border_style="thin")
medium_border = Side(border_style="medium")
def_font = Font(name='Arial', size=12)
def_border = Border(left=thin_border, right=thin_border, bottom=thin_border)
currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

def empty_row(row):
	for cell in row[1:6]:
		if is_full(cell):
			return False
	return True

def is_full(cell):
	return cell.value != None and str(cell.value) != ' ' and str(cell.value).lower() != 'off'

def import_dispatch_sheets(wb, sheets):
	for sheet in sheets:
		try:
			date = parser.parse(sheet)
			print('Valid date')
		except ValueError:
			print('Invalid date')
			continue
		app.config['LOADS'].remove({'date':date})
		parse_dispatch_sheet(wb[sheet], date, False)
		parse_dispatch_sheet(wb['Subhaul '+sheet], date, True)

def parse_dispatch_sheet(ws, date, subhaul):
	temp = None
	for row in ws.iter_rows(min_row=4, max_col=9):
		if is_full(row[0]) and 'total' in row[0].value.lower():
			app.config['LOADS'].insert(temp.__dict__)
			break
		elif is_full(row[0]) and not empty_row(row):
			if temp:
				app.config['LOADS'].insert(temp.__dict__)
			temp = Load(row, date, subhaul)
		elif temp and not empty_row(row):
			temp.set_other(row)

def create_payroll(date, filename):
	book = Workbook()
	sheet = book.active
	currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
	thin = Side(border_style="thin")
	medium = Side(border_style="medium")

	for i, driver in enumerate(app.config['DRIVERS'].find()):
		cell_info = [
			{
				'letter': 'A',
				'width': 20.5,
				'header': 'NAME',
				'key': 'driver'
			},{
				'letter': 'B',
				'width': 15.5,
				'header': 'HAND TAG',
				'key': 'hand_tag'
			},{
				'letter': 'C',
				'width': 24,
				'header': 'PICK-UP LOCATION',
				'key': 'origin'
			},{
				'letter': 'D',
				'width': 18,
				'header': 'CUSTOMER',
				'key': 'customer'
			},{
				'letter': 'E',
				'width': 18,
				'header': 'P.O. NUMBER',
				'key': 'po_num'
			},{
				'letter': 'F',
				'width': 20,
				'header': 'DESTINATION',
				'key': 'destination'
			},{
				'letter': 'G',
				'width': 15,
				'header': 'FORKLIFT',
				'key': 'forklift'
			},{
				'letter': 'H',
				'width': 21,
				'header': 'AMOUNT BILLED',
				'key': 'amount_billed'
			}
		] if driver['forklift'] else [
			{
				'letter': 'A',
				'width': 20.5,
				'header': 'NAME',
				'key': 'driver'
			},{
				'letter': 'B',
				'width': 15.5,
				'header': 'HAND TAG',
				'key': 'hand_tag'
			},{
				'letter': 'C',
				'width': 24,
				'header': 'PICK-UP LOCATION',
				'key': 'origin'
			},{
				'letter': 'D',
				'width': 18,
				'header': 'CUSTOMER',
				'key': 'customer'
			},{
				'letter': 'E',
				'width': 18,
				'header': 'P.O. NUMBER',
				'key': 'po_num'
			},{
				'letter': 'F',
				'width': 20,
				'header': 'DESTINATION',
				'key': 'destination'
			},{
				'letter': 'G',
				'width': 21,
				'header': 'AMOUNT BILLED',
				'key': 'amount_billed'
			}
		]
		sum_rows = ['G','H','I'] if driver['forklift'] else ['G','H']

		loads = list(app.config['LOADS'].find({'driver':driver['alias'],'date':{'$gte':date,'$lt':date+timedelta(days=5)}}))

		if not loads:
			continue

		if sheet.title in ['Sheet','Sheet1']:
			sheet.title = driver['name']
		else:
			sheet = book.create_sheet(driver['name'])

		sheet['A1'] = 'Driver Name:'
		sheet['B1'] = driver['name']
		sheet['A2'] = 'Week starting:'
		sheet['B2'] = date.strftime('%B %-d')
		sheet['A1'].font = Font(name='Arial', size=16, bold=True)
		sheet['A1'].border = Border(bottom=thin)
		sheet['B1'].font = Font(name='Arial', size=16)
		sheet['B1'].border = Border(bottom=thin)
		sheet['A2'].font = Font(name='Arial', size=14)
		sheet['B2'].font = Font(name='Arial', size=14)

		for i2,c in enumerate(cell_info + [{'letter':sum_rows[-1],'width':16,'header':str(driver['percentage'])+'%'}]):
			cell = sheet.cell(row=3, column=i2+1)
			sheet.column_dimensions[c['letter']].width = c['width']
			cell.value = c['header']
			cell.alignment = Alignment(horizontal='center')
			cell.font = Font(name='Arial', size=14, bold=True)
			cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)

		for i3,load in enumerate(loads):
			for i4, h in enumerate([a['key'] for a in cell_info]):
				cell = sheet.cell(row=i3+4, column=i4+1)
				cell.value = load[h]
				cell.font = Font(name='Arial', size=14)
				cell.border = Border(left=thin, right=thin, bottom=thin)
				if h in ['forklift','amount_billed']:
					cell.number_format = currency_format
				if h == 'forklift':
					cell.value = load[h]/3
			cell = sheet.cell(row=i3+4, column=len(cell_info)+1)
			cell.value = '='+sum_rows[-2]+str(i3+4)+'*0.'+str(driver['percentage'])
			cell.font = Font(name='Arial', size=14)
			cell.border = Border(left=thin, right=thin, bottom=thin)
			cell.number_format = currency_format

		tr = sheet.max_row
		for i5,letter in enumerate(sum_rows):
			cell = sheet.cell(row=tr+1, column=i5+7)
			cell.value = '=SUM('+letter+'4:'+letter+str(tr)+')'
			cell.font = Font(name='Arial', size=14)
			cell.number_format = currency_format

		final_sum = [('Gross load revenue:','='+sum_rows[-1]+str(tr+1)),('Standby Time (Hours)','=B26*50'),('Wide Load:',0),('Forklift:','=G'+str(tr+1)),('Total gross wage:','=SUM(C25:C28)')]

		for i6,a in enumerate(final_sum):
			if a[0] == 'Forklift:' and not driver['forklift']:
				continue
			cell1 = sheet.cell(row=i6+25, column=1)
			cell1.value = a[0]
			cell1.font = Font(name='Arial', size=12)
			cell2 = sheet.cell(row=i6+25, column=3)
			cell2.value = a[1]
			cell2.font = Font(name='Arial', size=12)
			cell2.number_format = currency_format

		sheet['B26'] = 0
		sheet['B26'].font = Font(name='Arial', size=12)
		sheet['C29'].border = Border(top=thin)

		sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
		sheet.sheet_properties.pageSetUpPr.fitToPage = True

	book.save(filename)

def create_price_sheet(origin,customer,data,filename):
	temp_customer = 'BMC'
	temp_address = 'Murphy Road, Salida'

	book = Workbook()
	sheet = book.active
	sheet.title = origin.split(', CA')[0]
	thin_border = Side(border_style="thin")
	medium_border = Side(border_style="medium")
	def_font = Font(name='Arial', size=12)
	def_border = Border(left=thin_border, right=thin_border, bottom=thin_border)
	miles_format = '#,##0.0'
	rate_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
	col_info = [
		{
			'letter': 'A',
			'width': 25,
			'header': 'CITY',
			'key': 'city'
		},{
			'letter': 'B',
			'width': 10,
			'header': 'MILES',
			'key': 'miles'
		},{
			'letter': 'C',
			'width': 20,
			'header': 'RATE',
			'key': 'rate'
		}
	]

	sheet['A1'] = 'Price sheet for '+customer
	sheet['A2'] = origin.split(', CA')[0]
	sheet['A1'].font = Font(name='Arial', size=14)
	sheet['A2'].font = Font(name='Arial', size=14)

	for i1,a in enumerate(col_info):
		cell = sheet.cell(row=3, column=1+i1)
		sheet.column_dimensions[a['letter']].width = a['width']
		cell.value = a['header']
		cell.alignment = Alignment(horizontal='center')
		cell.font = Font(name='Arial', size=12, bold=True)
		cell.border = Border(left=thin_border, right=thin_border, top=medium_border, bottom=medium_border)

	for i2,d in enumerate(data):
		for i3,c in enumerate(col_info):
			cell = sheet.cell(row=4+i2, column=1+i3)
			cell.value = d[c['key']]
			if c['key'] == 'miles':
				cell.number_format = miles_format
				cell.alignment = Alignment(horizontal='center')
			if c['key'] == 'rate':
				cell.value = '=ROUND((B'+str(4+i2)+'+5)/10,0)*25+300'
				cell.number_format = rate_format
			cell.font = def_font
			cell.border = def_border

	sheet = book.create_sheet('Sundance Mileage Chart')
	mileage_chart(sheet)
	book.save(filename)

def mileage_chart(sheet):
	img_dir = os.path.join(os.path.dirname(os.path.realpath(__file__)),'static/img/sundance_excel_header.png')
	img = Image(img_dir)
	# img.anchor('A1')
	sheet.add_image(img)

	for l,h in {'A':'MILEAGE','B':'RATE'}.items():
		sheet.column_dimensions[l].width = 15
		sheet[l+'9'].value = h
		sheet[l+'9'].alignment = Alignment(horizontal='center')
		sheet[l+'9'].font = Font(name='Arial', size=12, bold=True)
		sheet[l+'9'].border = Border(left=thin_border, right=thin_border, top=medium_border, bottom=medium_border)

	lb = 0
	n = 0
	for i in range(17):
		n = 20+(10*i)
		cell = sheet.cell(row=10+i, column=1)
		cell.value = str(lb)+'-'+str(n)
		cell.font = def_font
		cell.border = def_border
		lb = n+1
		cell = sheet.cell(row=10+i, column=2)
		cell.value = 300+(2.5*n)
		cell.font = def_font
		cell.border = def_border
		cell.number_format = currency_format

	notes = {
		'Forklift unload':'$75/drop',
		'Wide load':'$25',
		'Additional stop':'$50',
		'S/B time (after 1.5-hour loading)':'$110/hour',
		'S/B time (after 2-hour loading)':'$110/hour'
	}

	for i2,(note,price) in enumerate(notes.items()):
		cell = sheet.cell(row=28+i2, column=1)
		cell.value = note+' - '+price
		cell.font = def_font

	sheet['A34'].value = '*Rates effective as of ' + datetime.now().strftime('%b %-d, %Y')
	sheet['A34'].font = def_font










